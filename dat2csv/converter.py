import csv
import sys
from pathlib import Path


def _parse_dat(
    input_path: Path, encoding: str = "utf-8-sig"
) -> tuple[list[list[str]], int]:
    """
    Lê e parseia um arquivo .dat.

    Retorna (rows, max_cols) onde rows é a lista de campos por linha
    (já sem colunas vazias à direita) e max_cols é o maior número de
    colunas encontrado.
    """
    rows: list[list[str]] = []
    max_cols = 0

    with input_path.open(encoding=encoding) as f:
        for raw_line in f:
            raw_line = raw_line.strip()
            if not raw_line:
                continue

            reader = csv.reader([raw_line], quotechar="'", skipinitialspace=True)
            try:
                fields = next(reader)
            except StopIteration:
                continue

            while fields and fields[-1] == "":
                fields.pop()

            if fields:
                rows.append(fields)
                if len(fields) > max_cols:
                    max_cols = len(fields)

    return rows, max_cols


def _apply_value_labels(
    rows: list[list[str]],
    max_cols: int,
    value_labels: dict[str, dict[str, str]],
) -> list[list[str]]:
    """Substitui códigos pelos rótulos de valor nas colunas que possuem mapeamento."""
    resultado = []
    for row in rows:
        nova = list(row)
        for col_idx in range(len(nova)):
            var = f"V{col_idx + 1}"
            mapa = value_labels.get(var)
            if mapa and nova[col_idx] in mapa:
                nova[col_idx] = mapa[nova[col_idx]]
        resultado.append(nova)
    return resultado


def export_to_excel(
    registros: list[list[str]],
    cabecalhos: list[str] | None,
    arquivo_saida: str,
    backup: bool = True,
) -> dict:
    """
    Exporta registros para arquivo Excel (.xlsx).

    Args:
        registros:     Lista de linhas com os dados.
        cabecalhos:    Lista opcional com nomes das colunas.
        arquivo_saida: Caminho para o arquivo .xlsx.
        backup:        Se True (padrão), cria backup antes de sobrescrever.

    Returns:
        dict com chaves 'arquivo_saida', 'linhas', 'colunas', 'backup'.

    Raises:
        ImportError: Se openpyxl não estiver instalado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl não está instalado. Instale com 'pip install dat2csv[excel]'"
        )

    from .utils import criar_backup

    arquivo = Path(arquivo_saida)
    if arquivo.suffix.lower() != ".xlsx":
        arquivo = arquivo.with_suffix(".xlsx")

    arquivo.parent.mkdir(parents=True, exist_ok=True)
    backup_path = criar_backup(arquivo) if backup else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"

    num_cols = max(
        max((len(r) for r in registros), default=0),
        len(cabecalhos) if cabecalhos else 0,
    )

    # ── Cabeçalho ──────────────────────────────────────────────
    if cabecalhos:
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        for col_idx, nome in enumerate(cabecalhos, start=1):
            cell = ws.cell(row=1, column=col_idx, value=nome)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

    # ── Dados ──────────────────────────────────────────────────
    start_row = 2 if cabecalhos else 1
    for row_idx, row in enumerate(registros):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=start_row + row_idx, column=col_idx, value=valor)

    # ── Auto-ajuste de largura das colunas ─────────────────────
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        for row_cells in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, values_only=True
        ):
            for cell_val in row_cells:
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
        adjusted = min(max(8, max_len + 2), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    wb.save(arquivo)

    return {
        "arquivo_saida": str(arquivo),
        "linhas": len(registros),
        "colunas": num_cols,
        "backup": backup_path,
    }


def convert(
    input_path: str | Path,
    output_path: str | Path,
    encoding: str = "utf-8-sig",
    clean: bool = False,
    backup: bool = True,
    sps_path: str | Path | None = None,
    apply_labels: bool = False,
    add_header: bool = True,
    excel: bool = False,
) -> dict:
    """
    Converte um arquivo .dat para CSV ou Excel.

    Args:
        input_path:   Caminho para o arquivo .dat de entrada.
        output_path:  Caminho para o arquivo de saída (.csv ou .xlsx).
        encoding:     Encoding do arquivo de entrada (padrão: utf-8-sig).
        clean:        Se True, remove colunas 100% vazias do CSV/Excel final.
        backup:       Se True (padrão), cria backup do arquivo de saída caso já exista.
        sps_path:     Caminho opcional para arquivo .sps com metadados SPSS.
        apply_labels: Se True (requer sps_path), substitui códigos por rótulos de valor.
        add_header:   Se True (padrão) e sps_path fornecido, adiciona linha de cabeçalho.
        excel:        Se True, exporta para Excel (.xlsx) em vez de CSV. Requer openpyxl.

    Returns:
        dict com chaves 'rows', 'columns', 'backup' (Path ou None) e,
        se clean=True, 'removed_cols'.
    """
    # Importação local para evitar ciclo (utils importa converter)
    from .utils import criar_backup

    input_path = Path(input_path)
    output_path = Path(output_path)

    # ── Metadados do .sps ────────────────────────────────────────────────────
    metadata: dict = {"variable_labels": {}, "value_labels": {}}
    if sps_path is not None:
        from .sps import parse_sps
        metadata = parse_sps(sps_path)

    rows, max_cols = _parse_dat(input_path, encoding)

    # ── Substituição de labels ───────────────────────────────────────────────
    if apply_labels and metadata["value_labels"]:
        rows = _apply_value_labels(rows, max_cols, metadata["value_labels"])

    # ── Limpeza de colunas 100% vazias ───────────────────────────────────────
    keep_cols: list[int] | None = None
    removed = 0
    if clean and rows:
        empty = {
            col_idx
            for col_idx in range(max_cols)
            if all((col_idx >= len(r) or r[col_idx] == "") for r in rows)
        }
        keep_cols = [i for i in range(max_cols) if i not in empty]
        removed = len(empty)

    # ── Ajusta extensão para Excel ───────────────────────────────────────────
    if excel and output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = criar_backup(output_path) if backup else None

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    header: list[str] | None = None
    if sps_path is not None and add_header and metadata["variable_labels"]:
        all_headers = [
            metadata["variable_labels"].get(f"V{i + 1}", f"V{i + 1}")
            for i in range(max_cols)
        ]
        header = (
            [all_headers[i] for i in keep_cols]
            if keep_cols is not None
            else all_headers
        )

    # ── Escrita ──────────────────────────────────────────────────────────────
    if excel:
        # Monta registros normalizados para o Excel
        padded_rows = []
        for row in rows:
            padded = row + [""] * (max_cols - len(row))
            out_row = [padded[i] for i in keep_cols] if keep_cols is not None else padded
            padded_rows.append(out_row)

        result = export_to_excel(padded_rows, header, str(output_path), backup=False)
        final_cols = result["colunas"]
        result_dict: dict = {
            "rows": result["linhas"],
            "columns": final_cols,
            "backup": backup_path,
        }
        if clean:
            result_dict["removed_cols"] = removed
        return result_dict

    # ── Escrita do CSV ───────────────────────────────────────────────────────
    with output_path.open("w", encoding="utf-8", newline="") as f_out:
        writer = csv.writer(f_out)
        if header is not None:
            writer.writerow(header)
        for row in rows:
            padded = row + [""] * (max_cols - len(row))
            out_row = [padded[i] for i in keep_cols] if keep_cols is not None else padded
            writer.writerow(out_row)

    final_cols = len(keep_cols) if keep_cols is not None else max_cols
    result: dict = {"rows": len(rows), "columns": final_cols, "backup": backup_path}
    if clean:
        result["removed_cols"] = removed
    return result
