"""Testes para a funcionalidade de exportação para Excel (.xlsx)."""
import sys
from pathlib import Path
from unittest.mock import patch

import pytest

from dat2csv.converter import export_to_excel, convert


# ── Fixtures auxiliares ──────────────────────────────────────────────────────

REGISTROS_SIMPLES = [
    ["1", "dois", "3"],
    ["4", "cinco", "6"],
    ["7", "oito", "9"],
]

CABECALHOS_SIMPLES = ["id", "nome", "valor"]


# ── export_to_excel ──────────────────────────────────────────────────────────

class TestExportToExcel:
    """Testes da função export_to_excel."""

    def test_export_basico(self, tmp_path):
        """Exporta registros simples e verifica se arquivo .xlsx é criado."""
        saida = tmp_path / "saida.xlsx"
        result = export_to_excel(REGISTROS_SIMPLES, None, str(saida))

        assert saida.exists()
        assert result["linhas"] == 3
        assert result["colunas"] == 3
        assert result["backup"] is None

        # Verifica conteúdo com openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(saida)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "1"
        assert ws.cell(row=1, column=2).value == "dois"
        assert ws.cell(row=3, column=3).value == "9"

    def test_export_com_cabecalhos(self, tmp_path):
        """Com cabeçalhos: primeira linha em negrito, fundo cinza, freeze panes."""
        saida = tmp_path / "saida.xlsx"
        export_to_excel(REGISTROS_SIMPLES, CABECALHOS_SIMPLES, str(saida))

        import openpyxl
        wb = openpyxl.load_workbook(saida)
        ws = wb.active

        # Cabeçalho na linha 1
        assert ws.cell(row=1, column=1).value == "id"
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=1).fill.start_color.rgb == "00D9D9D9"

        # Freeze panes na A2
        assert ws.freeze_panes == "A2"

        # Dados começam na linha 2
        assert ws.cell(row=2, column=1).value == "1"

    def test_export_sem_cabecalhos(self, tmp_path):
        """Sem cabeçalhos, dados começam na linha 1 e sem freeze."""
        saida = tmp_path / "saida.xlsx"
        export_to_excel(REGISTROS_SIMPLES, None, str(saida))

        import openpyxl
        wb = openpyxl.load_workbook(saida)
        ws = wb.active

        # Dados começam na linha 1
        assert ws.cell(row=1, column=1).value == "1"
        # Sem freeze
        assert ws.freeze_panes is None or ws.freeze_panes == "A1"

    def test_export_com_backup(self, tmp_path):
        """Se arquivo já existe, cria backup antes de sobrescrever."""
        saida = tmp_path / "saida.xlsx"

        # Cria um arquivo .xlsx inicial
        export_to_excel([["X"]], None, str(saida))

        # Exporta novamente com backup=True
        result = export_to_excel(REGISTROS_SIMPLES, None, str(saida), backup=True)

        assert result["backup"] is not None
        assert result["backup"].exists()
        assert "_backup_" in result["backup"].name
        # O backup deve ter extensão .xlsx
        assert result["backup"].suffix == ".xlsx"

        # O arquivo original (saida.xlsx) deve conter os novos dados
        import openpyxl
        wb = openpyxl.load_workbook(saida)
        assert wb.active.cell(row=1, column=1).value == "1"

    def test_export_sem_backup(self, tmp_path):
        """Com backup=False, sobrescreve sem criar backup."""
        saida = tmp_path / "saida.xlsx"

        # Cria um arquivo .xlsx inicial
        export_to_excel([["X"]], None, str(saida))

        # Exporta novamente com backup=False
        result = export_to_excel(REGISTROS_SIMPLES, None, str(saida), backup=False)

        assert result["backup"] is None
        # Nenhum arquivo _backup_ no diretório
        backups = list(tmp_path.glob("*_backup_*"))
        assert backups == []

    def test_export_extensao_automatica(self, tmp_path):
        """Se nome não termina com .xlsx, adiciona automaticamente."""
        saida = tmp_path / "saida"  # sem extensão
        result = export_to_excel(REGISTROS_SIMPLES, None, str(saida))

        arquivo = Path(result["arquivo_saida"])
        assert arquivo.suffix == ".xlsx"
        assert arquivo.exists()

        # Também testa com .csv
        saida_csv = tmp_path / "saida.csv"
        result2 = export_to_excel(REGISTROS_SIMPLES, None, str(saida_csv))
        arquivo2 = Path(result2["arquivo_saida"])
        assert arquivo2.suffix == ".xlsx"
        assert arquivo2.exists()

    def test_export_auto_ajuste_largura(self, tmp_path):
        """Colunas devem ter largura ajustada automaticamente."""
        saida = tmp_path / "saida.xlsx"
        registros = [["coluna_com_texto_longo"], ["curto"]]
        export_to_excel(registros, ["cabecalho_bastante_longo"], str(saida))

        import openpyxl
        wb = openpyxl.load_workbook(saida)
        ws = wb.active

        # A largura deve ser > 8 (mínimo) e ≤ 50 (máximo)
        width = ws.column_dimensions["A"].width
        assert width >= 8
        assert width <= 50

    def test_export_planilha_titulo_dados(self, tmp_path):
        """A planilha deve ter o título 'Dados'."""
        saida = tmp_path / "saida.xlsx"
        export_to_excel(REGISTROS_SIMPLES, None, str(saida))

        import openpyxl
        wb = openpyxl.load_workbook(saida)
        assert wb.active.title == "Dados"

    def test_export_falha_openpyxl(self, tmp_path):
        """Se openpyxl não instalado, levanta ImportError com mensagem clara."""
        saida = tmp_path / "saida.xlsx"

        # Simula openpyxl ausente mockando sys.modules
        with patch.dict(sys.modules, {"openpyxl": None, "openpyxl.styles": None,
                                       "openpyxl.utils": None}):
            with pytest.raises(ImportError, match="openpyxl não está instalado"):
                export_to_excel(REGISTROS_SIMPLES, None, str(saida))


# ── convert() com excel=True ─────────────────────────────────────────────────

class TestConvertExcel:
    """Testes de integração: convert() com excel=True."""

    def test_convert_basico_excel(self, dat_simples, tmp_path):
        """convert() com excel=True gera .xlsx."""
        out = tmp_path / "out.csv"  # extensão .csv será ajustada
        result = convert(dat_simples, out, backup=False, excel=True)

        # Deve ter ajustado a extensão para .xlsx
        saida_real = out.with_suffix(".xlsx")
        assert saida_real.exists()
        assert result["rows"] == 3
        assert result["columns"] == 3

    def test_convert_excel_com_sps(self, tmp_path, sps_simples):
        """convert() com excel=True + sps_path gera .xlsx com cabeçalhos."""
        dat = tmp_path / "d.dat"
        dat.write_text("1,dois,3\n4,cinco,6\n", encoding="utf-8")
        out = tmp_path / "out.xlsx"
        result = convert(dat, out, backup=False, sps_path=sps_simples, excel=True)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "id"
        assert ws.cell(row=1, column=2).value == "nome"
        assert ws.cell(row=1, column=3).value == "genero"
        assert ws.cell(row=2, column=1).value == "1"

    def test_convert_excel_com_clean(self, dat_colunas_vazias, tmp_path):
        """convert() com excel=True e clean=True remove colunas vazias."""
        out = tmp_path / "out.xlsx"
        result = convert(dat_colunas_vazias, out, backup=False, clean=True, excel=True)

        assert result["removed_cols"] == 1
        assert result["columns"] == 3

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # Deve ter 3 colunas (a vazia foi removida)
        assert ws.max_column == 3

    def test_convert_excel_no_header(self, tmp_path, sps_simples):
        """convert() com excel=True e add_header=False não escreve cabeçalho."""
        dat = tmp_path / "d.dat"
        dat.write_text("1,dois,3\n", encoding="utf-8")
        out = tmp_path / "out.xlsx"
        convert(dat, out, backup=False, sps_path=sps_simples,
                add_header=False, excel=True)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "1"

    def test_convert_excel_apply_labels(self, tmp_path, sps_com_labels):
        """convert() com excel=True e apply_labels=True substitui códigos."""
        dat = tmp_path / "d.dat"
        dat.write_text("1,AO01,AO01\n", encoding="utf-8")
        out = tmp_path / "out.xlsx"
        convert(dat, out, backup=False, sps_path=sps_com_labels,
                apply_labels=True, excel=True)

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "Portugal"


# ── CLI: flag --excel ────────────────────────────────────────────────────────

class TestCliExcel:
    """Testes da flag --excel na interface de linha de comando."""

    def test_cli_flag_excel(self, dat_simples, tmp_path, capsys):
        """--excel gera arquivo .xlsx."""
        from dat2csv.cli import main

        out = tmp_path / "out.xlsx"
        sys.argv = ["dat2csv", str(dat_simples), "--excel",
                    "--excel-output", str(out)]
        main()
        assert out.exists()

        captured = capsys.readouterr()
        assert "Arquivo convertido com sucesso" in captured.out

    def test_cli_excel_output_padrao(self, dat_simples, tmp_path, capsys):
        """--excel sem --excel-output usa nome do .dat com .xlsx."""
        from dat2csv.cli import main

        # Copia o .dat para um diretório temporário para controlar a saída
        dat = tmp_path / "dados.dat"
        dat.write_text(dat_simples.read_text(), encoding="utf-8")

        sys.argv = ["dat2csv", str(dat), "--excel"]
        main()

        saida = tmp_path / "dados.xlsx"
        assert saida.exists()

    def test_cli_excel_output_customizado(self, dat_simples, tmp_path, capsys):
        """--excel-output define caminho personalizado."""
        from dat2csv.cli import main

        out = tmp_path / "personalizado.xlsx"
        sys.argv = ["dat2csv", str(dat_simples), "--excel",
                    "--excel-output", str(out)]
        main()
        assert out.exists()

    def test_cli_excel_output_ignorado_sem_excel(self, dat_simples, tmp_path, capsys):
        """--excel-output sem --excel gera aviso e converte para CSV."""
        from dat2csv.cli import main

        out = tmp_path / "out.csv"
        sys.argv = ["dat2csv", str(dat_simples),
                    "--excel-output", str(tmp_path / "ignorado.xlsx"),
                    str(out)]
        main()
        captured = capsys.readouterr()
        assert "Aviso" in captured.err
        assert out.exists()  # gerou CSV, não Excel

    def test_cli_excel_incompativel_preview(self, dat_simples, capsys):
        """--excel --preview gera erro."""
        from dat2csv.cli import main

        sys.argv = ["dat2csv", str(dat_simples), "--excel", "--preview"]
        with pytest.raises(SystemExit):
            main()
        captured = capsys.readouterr()
        assert "Erro: --excel não pode ser usado junto com --preview" in captured.err

    def test_cli_excel_incompativel_inspect(self, dat_simples, capsys):
        """--excel --inspect gera erro."""
        from dat2csv.cli import main

        sys.argv = ["dat2csv", str(dat_simples), "--excel", "--inspect"]
        with pytest.raises(SystemExit):
            main()
        captured = capsys.readouterr()
        assert "Erro: --excel não pode ser usado junto com --inspect" in captured.err

    def test_cli_excel_ignora_output(self, dat_simples, tmp_path, capsys):
        """--excel com --output: ignora --output com aviso."""
        from dat2csv.cli import main

        out_csv = tmp_path / "ignorado.csv"
        out_xlsx = tmp_path / "dados.xlsx"
        dat = tmp_path / "dados.dat"
        dat.write_text(dat_simples.read_text(), encoding="utf-8")

        sys.argv = ["dat2csv", str(dat), "--excel", str(out_csv)]
        main()
        captured = capsys.readouterr()
        assert "Aviso" in captured.err
        assert "ignorado" in captured.err
        # Deve ter gerado .xlsx (padrão), não o .csv
        assert out_xlsx.exists()
        assert not out_csv.exists()

    def test_cli_excel_com_sps(self, tmp_path, sps_simples, capsys):
        """--excel com --sps gera .xlsx com cabeçalhos."""
        from dat2csv.cli import main

        dat = tmp_path / "d.dat"
        dat.write_text("1,dois,3\n", encoding="utf-8")
        out = tmp_path / "out.xlsx"

        sys.argv = ["dat2csv", str(dat), "--sps", str(sps_simples),
                    "--excel", "--excel-output", str(out)]
        main()

        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "id"
